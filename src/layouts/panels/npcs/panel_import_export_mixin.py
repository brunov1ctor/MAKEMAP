"""ImportExportMixin — the Importar (JSON/CSV paste cards + Excel drop
card) / Exportar (read-only JSON/CSV view + direct-to-file Excel) tools
panel that takes over the right column while active. Mixed into NPCsPanel
(see panel.py) — operates on self.* attributes NPCsPanel owns; not meant to
be instantiated on its own.

Mirrors src/layouts/panels/mobs/panel_import_export_mixin.py.
"""

from __future__ import annotations

import logging

from PySide6.QtWidgets import (
    QFrame, QVBoxLayout, QHBoxLayout, QLabel, QTextEdit,
    QPushButton, QWidget, QStackedWidget, QScrollArea, QFileDialog,
)
from PySide6.QtCore import Qt

from src.styles.tokens import Colors
from src.layouts.panels.npcs.npc_edit_panel import NPCEditPanel
from src.layouts.panels.npcs.panel_widgets import _DropZone
from src.layouts.panels.npcs.panel_helpers import (
    _NPC_TEMPLATE_FIELDS, _TEMPLATE_FIELD_DOCS, _parse_npcs_json, _normalize_npc_name,
)
from src.layouts.panels.shared.import_export_helpers import (
    index_files_by_stem, normalize_blank_cells, read_json, read_csv, read_xlsx,
    import_button_row, build_tools_header, build_export_view_page,
)
from src.services.project_assets import import_asset

_IMAGE_EXTS = (".png", ".jpg", ".jpeg", ".webp", ".bmp")
_ASSET_EXTS = (".fbx", ".obj", ".gltf", ".glb", ".png", ".jpg", ".jpeg", ".webp")

logger = logging.getLogger("MAKEMAP")


class ImportExportMixin:
    """The right column's alternate page (see self._right_stack) — a
    header (title + ✕ close) above a small stack of its own: page 0 is
    Importar — separate JSON / CSV (paste text) / Excel (drop file) cards,
    each with its own Cancelar/Template/Aplicar row — page 1 is a
    read-only JSON/CSV export view of the user's current npcs (Exportar)."""

    def _build_tools_panel(self) -> QWidget:
        panel = QFrame()
        panel.setStyleSheet(f"QFrame {{ background: rgba(255,255,255,0.03); border-radius: 8px; }}")
        outer = QVBoxLayout(panel)
        outer.setContentsMargins(10, 10, 10, 10)
        outer.setSpacing(8)

        head_row, self._tools_title_lbl, _tools_close_btn = build_tools_header(
            self._close_tools_mode, title_font_size=12, close_font_size=12)
        outer.addLayout(head_row)

        self._tools_stack = QStackedWidget()

        import_content = QWidget()
        import_lay = QVBoxLayout(import_content)
        import_lay.setContentsMargins(0, 0, 0, 0)
        import_lay.setSpacing(8)
        import_lay.addWidget(self._build_text_import_card(
            "JSON", "Cole uma lista JSON para criar vários npcs de uma vez.", "json"))
        import_lay.addWidget(self._build_text_import_card(
            "CSV", "Cole uma lista CSV para criar vários npcs de uma vez.", "csv"))
        import_lay.addWidget(self._build_excel_import_card())
        import_lay.addWidget(self._build_image_folder_import_card())
        import_lay.addWidget(self._build_asset_folder_import_card())
        import_lay.addStretch()
        import_scroll = QScrollArea()
        import_scroll.setWidgetResizable(True)
        import_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        import_scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }"
                                     "QScrollArea > QWidget > QWidget { background: transparent; }")
        import_scroll.setWidget(import_content)
        self._tools_stack.addWidget(import_scroll)

        template_page, self._template_hint_lbl, self._template_edit = build_export_view_page(self._on_save_export_file)
        self._tools_stack.addWidget(template_page)
        outer.addWidget(self._tools_stack, 1)

        panel.setMinimumWidth(NPCEditPanel.PANEL_WIDTH)
        return panel

    # ─── Importar cards (JSON/CSV paste, Excel drop) ───

    def _import_button_row(self, on_cancel, on_template, on_apply) -> tuple[QHBoxLayout, QPushButton]:
        return import_button_row(on_cancel, on_template, on_apply)

    def _json_import_template(self) -> str:
        import json
        valid_categories = ", ".join(
            f"{c['id']} ({c['name']})" for c in sorted(self._all_categories, key=lambda c: c["name"])
        ) or "nenhuma ainda — crie uma no explorador à esquerda"
        doc_lines = "\n".join(f"// {key}: {doc}" for key, doc in _TEMPLATE_FIELD_DOCS)
        cat_line = f"// Categorias disponíveis: {valid_categories}"
        example = dict(_NPC_TEMPLATE_FIELDS)
        example["name"] = "Novo NPC"
        row_json = json.dumps(example, ensure_ascii=False, indent=2)
        row_indented = "\n".join(f"  {line}" for line in row_json.splitlines())
        return f"{doc_lines}\n{cat_line}\n[\n{row_indented}\n]"

    def _csv_import_template(self) -> str:
        import csv
        import io
        example = dict(_NPC_TEMPLATE_FIELDS)
        example["name"] = "Novo NPC"
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=list(_NPC_TEMPLATE_FIELDS.keys()))
        writer.writeheader()
        writer.writerow(example)
        return buf.getvalue()

    def _build_text_import_card(self, title: str, hint: str, fmt: str) -> QFrame:
        card = QFrame()
        card.setStyleSheet(f"QFrame {{ background: rgba(0,0,0,0.15); border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 8px; }}")
        lay = QVBoxLayout(card)
        lay.setContentsMargins(10, 8, 10, 8)
        lay.setSpacing(5)

        title_lbl = QLabel(title)
        title_lbl.setStyleSheet(f"color: {Colors.TEXT_PRIMARY}; font-size: 11px; font-weight: bold; background: transparent; border: none;")
        lay.addWidget(title_lbl)
        hint_lbl = QLabel(hint)
        hint_lbl.setWordWrap(True)
        hint_lbl.setStyleSheet(f"color: {Colors.TEXT_MUTED}; font-size: 9px; background: transparent; border: none;")
        lay.addWidget(hint_lbl)

        edit = QTextEdit()
        edit.setFixedHeight(120)
        edit.setStyleSheet(f"""
            QTextEdit {{ color: {Colors.TEXT_PRIMARY}; font-size: 9px; font-family: Consolas, monospace;
                background: rgba(0,0,0,0.25); border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 4px; padding: 5px; }}
        """)
        lay.addWidget(edit)

        error_lbl = QLabel("")
        error_lbl.setWordWrap(True)
        error_lbl.setStyleSheet(f"color: {Colors.ERROR}; font-size: 9px; background: transparent; border: none;")
        error_lbl.hide()
        lay.addWidget(error_lbl)

        def reset_to_template():
            edit.setPlainText(self._json_import_template() if fmt == "json" else self._csv_import_template())
            error_lbl.hide()

        def do_apply():
            text = edit.toPlainText()
            try:
                if fmt == "json":
                    data = _parse_npcs_json(text)
                else:
                    import csv
                    import io
                    data = [self._normalize_blank_cells(dict(row)) for row in csv.DictReader(io.StringIO(text))]
            except ValueError as exc:
                error_lbl.setText(str(exc))
                error_lbl.show()
                return
            except Exception:
                logger.exception("Falha ao interpretar import de npcs (%s).", fmt)
                error_lbl.setText("Não foi possível interpretar o conteúdo — confira o formato.")
                error_lbl.show()
                return
            imported = self._import_npc_dicts(data)
            if imported == 0:
                error_lbl.setText('Nenhum npc válido encontrado (o campo "name" é obrigatório).')
                error_lbl.show()
                return
            logger.info("Criados %d npc(s) via Importar (%s)", imported, fmt)
            self._reload()
            reset_to_template()
            self._close_tools_mode()

        row, _apply_btn = self._import_button_row(self._close_tools_mode, reset_to_template, do_apply)
        lay.addLayout(row)
        reset_to_template()
        return card

    def _build_excel_import_card(self) -> QFrame:
        card = QFrame()
        card.setStyleSheet(f"QFrame {{ background: rgba(0,0,0,0.15); border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 8px; }}")
        lay = QVBoxLayout(card)
        lay.setContentsMargins(10, 8, 10, 8)
        lay.setSpacing(5)

        title_lbl = QLabel("Excel")
        title_lbl.setStyleSheet(f"color: {Colors.TEXT_PRIMARY}; font-size: 11px; font-weight: bold; background: transparent; border: none;")
        lay.addWidget(title_lbl)

        drop_zone = _DropZone()
        lay.addWidget(drop_zone)

        staged_lbl = QLabel("Nenhum arquivo selecionado.")
        staged_lbl.setWordWrap(True)
        staged_lbl.setStyleSheet(f"color: {Colors.TEXT_MUTED}; font-size: 9px; background: transparent; border: none;")
        lay.addWidget(staged_lbl)

        state = {"path": ""}

        def on_file_staged(path: str):
            state["path"] = path
            import os
            staged_lbl.setText(f"Selecionado: {os.path.basename(path)}")
            apply_btn.setEnabled(True)

        drop_zone.file_chosen.connect(on_file_staged)

        def clear_staged():
            state["path"] = ""
            staged_lbl.setText("Nenhum arquivo selecionado.")
            apply_btn.setEnabled(False)

        def do_cancel():
            clear_staged()
            self._close_tools_mode()

        def do_apply():
            if not state["path"]:
                return
            self._on_file_dropped(state["path"])
            clear_staged()

        row, apply_btn = self._import_button_row(do_cancel, self._export_xlsx_blank_template, do_apply)
        apply_btn.setEnabled(False)
        lay.addLayout(row)
        return card

    def _build_image_folder_import_card(self) -> QFrame:
        card = QFrame()
        card.setStyleSheet(f"QFrame {{ background: rgba(0,0,0,0.15); border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 8px; }}")
        lay = QVBoxLayout(card)
        lay.setContentsMargins(10, 8, 10, 8)
        lay.setSpacing(5)

        title_lbl = QLabel("Imagens (pasta)")
        title_lbl.setStyleSheet(f"color: {Colors.TEXT_PRIMARY}; font-size: 11px; font-weight: bold; background: transparent; border: none;")
        lay.addWidget(title_lbl)

        hint_lbl = QLabel(
            "Escolha uma pasta com imagens nomeadas como os npcs (ex.: "
            "\"Ferreiro da Vila.png\") — cada arquivo cujo nome bater com "
            "um npc existente recebe essa imagem como retrato."
        )
        hint_lbl.setWordWrap(True)
        hint_lbl.setStyleSheet(f"color: {Colors.TEXT_MUTED}; font-size: 9px; background: transparent; border: none;")
        lay.addWidget(hint_lbl)

        pick_btn = QPushButton("📁 Selecionar Pasta")
        pick_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        pick_btn.setStyleSheet(f"""
            QPushButton {{ background: rgba(255,255,255,0.06); color: {Colors.TEXT_PRIMARY};
                border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 6px; padding: 5px 10px; font-size: 10px; }}
            QPushButton:hover {{ background: rgba(255,255,255,0.12); border-color: {Colors.ACCENT}; }}
        """)
        lay.addWidget(pick_btn)

        preview_edit = QTextEdit()
        preview_edit.setReadOnly(True)
        preview_edit.setFixedHeight(120)
        preview_edit.setStyleSheet(f"""
            QTextEdit {{ color: {Colors.TEXT_PRIMARY}; font-size: 9px; font-family: Consolas, monospace;
                background: rgba(0,0,0,0.2); border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 4px; padding: 6px; }}
        """)
        preview_edit.setPlainText("Nenhuma pasta selecionada.")
        lay.addWidget(preview_edit)

        state = {"matches": []}  # list of (npc_dict, source_file_path)

        def refresh_preview():
            matches, unmatched = self._match_npcs_against_staged_folder()
            state["matches"] = matches
            preview_edit.setPlainText(self._format_image_match_preview(matches, unmatched))
            apply_btn.setEnabled(bool(matches))

        def on_pick_folder():
            folder = QFileDialog.getExistingDirectory(self, "Selecione a pasta com as imagens dos npcs")
            if not folder:
                return
            self._staged_image_folder = folder
            self._staged_image_files = self._index_image_folder(folder)
            refresh_preview()

        pick_btn.clicked.connect(on_pick_folder)

        def clear_staged():
            state["matches"] = []
            preview_edit.setPlainText("Nenhuma pasta selecionada.")
            apply_btn.setEnabled(False)

        def do_cancel():
            state["matches"] = []
            self._close_tools_mode()

        def do_apply():
            if not state["matches"]:
                return
            self._apply_image_matches(state["matches"])
            clear_staged()

        row = QHBoxLayout()
        row.addStretch()
        cancel_btn = QPushButton("Cancelar")
        cancel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        cancel_btn.setStyleSheet(f"""
            QPushButton {{ background: rgba(255,255,255,0.06); color: {Colors.TEXT_SECONDARY}; border: none;
                border-radius: 6px; padding: 5px 10px; font-size: 10px; }}
            QPushButton:hover {{ background: rgba(255,255,255,0.12); }}
        """)
        cancel_btn.clicked.connect(do_cancel)
        row.addWidget(cancel_btn)
        apply_btn = QPushButton("Aplicar")
        apply_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        apply_btn.setStyleSheet(f"""
            QPushButton {{ background: {Colors.ACCENT}; color: #08131F; border: none;
                border-radius: 6px; padding: 5px 12px; font-size: 10px; font-weight: bold; }}
            QPushButton:hover:enabled {{ background: {Colors.ACCENT_HOVER}; }}
            QPushButton:disabled {{ background: rgba(255,255,255,0.08); color: {Colors.TEXT_MUTED}; }}
        """)
        apply_btn.setEnabled(False)
        apply_btn.clicked.connect(do_apply)
        row.addWidget(apply_btn)
        lay.addLayout(row)
        return card

    @staticmethod
    def _index_image_folder(folder: str) -> dict[str, str]:
        return index_files_by_stem(folder, _IMAGE_EXTS)

    def _match_npcs_against_staged_folder(self) -> tuple[list, list[str]]:
        import os
        files_by_name = self._staged_image_files
        matches = []
        matched_keys = set()
        for m in self._npcs:
            key = _normalize_npc_name(m.get("name", ""))
            if key and key in files_by_name:
                matches.append((m, files_by_name[key]))
                matched_keys.add(key)

        unmatched = [
            os.path.basename(path) for key, path in files_by_name.items()
            if key not in matched_keys
        ]
        return matches, sorted(unmatched)

    @staticmethod
    def _format_image_match_preview(matches: list, unmatched: list[str]) -> str:
        import os
        lines = [f"{len(matches) + len(unmatched)} imagem(ns) encontradas, {len(matches)} npc(s) correspondente(s):"]
        for npc, path in matches:
            suffix = " (substituirá imagem atual)" if npc.get("image_path") else ""
            lines.append(f"  OK {npc.get('name', '')}  <-  {os.path.basename(path)}{suffix}")
        if unmatched:
            lines.append(f"{len(unmatched)} arquivo(s) sem npc correspondente:")
            for name in unmatched:
                lines.append(f"  - {name}")
        return "\n".join(lines)

    def _apply_image_matches(self, matches: list):
        if not self._uow:
            return
        applied = 0
        for npc, source_path in matches:
            npc_id = npc.get("id")
            if not npc_id:
                continue
            try:
                rel_path = import_asset(self._project_dir, source_path, "assets/npcs", npc_id)
                self._uow.npcs.update(npc_id, image_path=rel_path)
                applied += 1
            except Exception:
                logger.exception("Falha ao atribuir imagem ao npc %s (%s)", npc.get("name", ""), source_path)
        logger.info("Imagens atribuídas via pasta: %d npc(s)", applied)
        self._reload()
        self._close_tools_mode()

    # ─── Assets (pasta) — bulk npc_assets import, mirrors the image-folder
    # card above. ───

    def _build_asset_folder_import_card(self) -> QFrame:
        card = QFrame()
        card.setStyleSheet(f"QFrame {{ background: rgba(0,0,0,0.15); border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 8px; }}")
        lay = QVBoxLayout(card)
        lay.setContentsMargins(10, 8, 10, 8)
        lay.setSpacing(5)

        title_lbl = QLabel("Assets (pasta)")
        title_lbl.setStyleSheet(f"color: {Colors.TEXT_PRIMARY}; font-size: 11px; font-weight: bold; background: transparent; border: none;")
        lay.addWidget(title_lbl)

        hint_lbl = QLabel(
            "Escolha uma pasta com modelos/texturas cujo nome CONTENHA o "
            "nome de um npc (ex.: \"tex_ferreiro_01.jpg\", \"model_ferreiro.fbx\") "
            "— cada arquivo é adicionado como um novo asset do npc "
            "correspondente; um npc pode receber vários arquivos de uma vez."
        )
        hint_lbl.setWordWrap(True)
        hint_lbl.setStyleSheet(f"color: {Colors.TEXT_MUTED}; font-size: 9px; background: transparent; border: none;")
        lay.addWidget(hint_lbl)

        pick_btn = QPushButton("📁 Selecionar Pasta")
        pick_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        pick_btn.setStyleSheet(f"""
            QPushButton {{ background: rgba(255,255,255,0.06); color: {Colors.TEXT_PRIMARY};
                border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 6px; padding: 5px 10px; font-size: 10px; }}
            QPushButton:hover {{ background: rgba(255,255,255,0.12); border-color: {Colors.ACCENT}; }}
        """)
        lay.addWidget(pick_btn)

        preview_edit = QTextEdit()
        preview_edit.setReadOnly(True)
        preview_edit.setFixedHeight(120)
        preview_edit.setStyleSheet(f"""
            QTextEdit {{ color: {Colors.TEXT_PRIMARY}; font-size: 9px; font-family: Consolas, monospace;
                background: rgba(0,0,0,0.2); border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 4px; padding: 6px; }}
        """)
        preview_edit.setPlainText("Nenhuma pasta selecionada.")
        lay.addWidget(preview_edit)

        state = {"matches": []}  # list of (npc_dict, [source_file_paths])

        def refresh_preview():
            matches, unmatched = self._match_npcs_against_staged_asset_folder()
            state["matches"] = matches
            preview_edit.setPlainText(self._format_asset_match_preview(matches, unmatched))
            apply_btn.setEnabled(bool(matches))

        def on_pick_folder():
            folder = QFileDialog.getExistingDirectory(self, "Selecione a pasta com os assets dos npcs")
            if not folder:
                return
            self._staged_asset_folder = folder
            self._staged_asset_files = self._index_asset_folder(folder)
            refresh_preview()

        pick_btn.clicked.connect(on_pick_folder)

        def clear_staged():
            state["matches"] = []
            preview_edit.setPlainText("Nenhuma pasta selecionada.")
            apply_btn.setEnabled(False)

        def do_cancel():
            state["matches"] = []
            self._close_tools_mode()

        def do_apply():
            if not state["matches"]:
                return
            self._apply_asset_matches(state["matches"])
            clear_staged()

        row = QHBoxLayout()
        row.addStretch()
        cancel_btn = QPushButton("Cancelar")
        cancel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        cancel_btn.setStyleSheet(f"""
            QPushButton {{ background: rgba(255,255,255,0.06); color: {Colors.TEXT_SECONDARY}; border: none;
                border-radius: 6px; padding: 5px 10px; font-size: 10px; }}
            QPushButton:hover {{ background: rgba(255,255,255,0.12); }}
        """)
        cancel_btn.clicked.connect(do_cancel)
        row.addWidget(cancel_btn)
        apply_btn = QPushButton("Aplicar")
        apply_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        apply_btn.setStyleSheet(f"""
            QPushButton {{ background: {Colors.ACCENT}; color: #08131F; border: none;
                border-radius: 6px; padding: 5px 12px; font-size: 10px; font-weight: bold; }}
            QPushButton:hover:enabled {{ background: {Colors.ACCENT_HOVER}; }}
            QPushButton:disabled {{ background: rgba(255,255,255,0.08); color: {Colors.TEXT_MUTED}; }}
        """)
        apply_btn.setEnabled(False)
        apply_btn.clicked.connect(do_apply)
        row.addWidget(apply_btn)
        lay.addLayout(row)
        return card

    @staticmethod
    def _index_asset_folder(folder: str) -> dict[str, list[str]]:
        import os
        try:
            entries = os.listdir(folder)
        except OSError:
            logger.exception("Falha ao listar pasta de assets: %s", folder)
            return {}
        files_by_name: dict[str, list[str]] = {}
        for entry in entries:
            path = os.path.join(folder, entry)
            stem, ext = os.path.splitext(entry)
            if ext.lower() in _ASSET_EXTS and os.path.isfile(path):
                files_by_name.setdefault(_normalize_npc_name(stem), []).append(path)
        return files_by_name

    def _match_npcs_against_staged_asset_folder(self) -> tuple[list, list[str]]:
        import os
        files_by_name = self._staged_asset_files
        matches: list[tuple[dict, list[str]]] = []
        matched_paths: set[str] = set()
        for m in self._npcs:
            key = _normalize_npc_name(m.get("name", ""))
            if not key:
                continue
            npc_files = sorted(
                path for norm_stem, paths in files_by_name.items() if key in norm_stem for path in paths
            )
            if npc_files:
                matches.append((m, npc_files))
                matched_paths.update(npc_files)

        unmatched = [
            os.path.basename(path) for paths in files_by_name.values() for path in paths
            if path not in matched_paths
        ]
        return matches, sorted(unmatched)

    @staticmethod
    def _format_asset_match_preview(matches: list, unmatched: list[str]) -> str:
        import os
        total_files = sum(len(paths) for _m, paths in matches) + len(unmatched)
        lines = [f"{total_files} arquivo(s) encontrados, {len(matches)} npc(s) correspondente(s):"]
        for npc, paths in matches:
            lines.append(f"  OK {npc.get('name', '')}  ({len(paths)} arquivo(s)):")
            for path in paths:
                lines.append(f"      - {os.path.basename(path)}")
        if unmatched:
            lines.append(f"{len(unmatched)} arquivo(s) sem npc correspondente:")
            for name in unmatched:
                lines.append(f"  - {name}")
        return "\n".join(lines)

    def _apply_asset_matches(self, matches: list):
        if not self._uow:
            return
        import os
        import uuid
        applied = 0
        affected_npc_ids = set()
        for npc, paths in matches:
            npc_id = npc.get("id")
            if not npc_id:
                continue
            for path in paths:
                try:
                    asset_id = str(uuid.uuid4())
                    ext = os.path.splitext(path)[1].lower()
                    asset_type = {
                        ".fbx": "Modelo 3D", ".obj": "Modelo 3D", ".gltf": "Modelo 3D", ".glb": "Modelo 3D",
                        ".png": "Imagem", ".jpg": "Imagem", ".jpeg": "Imagem", ".webp": "Imagem",
                    }.get(ext, "Arquivo")
                    try:
                        size = os.path.getsize(path)
                    except OSError:
                        size = 0
                    rel_path = import_asset(self._project_dir, path, "assets/npc_assets", asset_id)
                    self._uow.npc_assets.create(
                        id=asset_id, npc_id=npc_id, name=os.path.basename(path),
                        asset_type=asset_type, file_path=rel_path, file_size=size, rarity="common",
                    )
                    applied += 1
                    affected_npc_ids.add(npc_id)
                except Exception:
                    logger.exception("Falha ao importar asset para o npc %s (%s)", npc.get("name", ""), path)
        logger.info("Assets importados via pasta: %d arquivo(s) em %d npc(s)", applied, len(affected_npc_ids))
        if self._selected_id in affected_npc_ids:
            self._refresh_assets_display(self._selected_id)
        self._reload()
        self._close_tools_mode()

    def _toggle_import_mode(self):
        if self._tools_mode == "import":
            self._close_tools_mode()
            return
        self._tools_title_lbl.setText("Importar NPCs")
        self._tools_stack.setCurrentIndex(0)
        self._right_stack.setCurrentIndex(1)
        self._tools_mode = "import"

    def _on_export_choice(self, fmt: str):
        if fmt == "xlsx":
            self._export_xlsx()
            return
        mode = f"export_{fmt}"
        if self._tools_mode == mode:
            self._close_tools_mode()
            return
        self._template_fmt = fmt
        self._template_edit.setPlainText(self._build_npc_export(fmt))
        self._template_hint_lbl.setText(
            f"{len(self._npcs)} npc(s) — clique em Salvar Arquivo para exportar, ou copie o texto abaixo."
        )
        self._tools_title_lbl.setText("Exportar como JSON" if fmt == "json" else "Exportar como CSV")
        self._tools_stack.setCurrentIndex(1)
        self._right_stack.setCurrentIndex(1)
        self._tools_mode = mode

    def _close_tools_mode(self):
        self._right_stack.setCurrentIndex(0)
        self._tools_mode = None

    def _build_npc_export(self, fmt: str) -> str:
        fieldnames = list(_NPC_TEMPLATE_FIELDS.keys())
        rows = [{k: m.get(k, _NPC_TEMPLATE_FIELDS[k]) for k in fieldnames} for m in self._npcs]
        if fmt == "json":
            import json
            return json.dumps(rows, ensure_ascii=False, indent=2)
        import csv
        import io
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
        return buf.getvalue()

    def _on_save_export_file(self):
        fmt = self._template_fmt
        ext = "json" if fmt == "json" else "csv"
        file_filter = "JSON (*.json)" if fmt == "json" else "CSV (*.csv)"
        path, _selected = QFileDialog.getSaveFileName(self, "Salvar Exportação de NPCs", f"npcs.{ext}", file_filter)
        if not path:
            return
        if not path.lower().endswith(f".{ext}"):
            path += f".{ext}"
        with open(path, "w", encoding="utf-8", newline="") as f:
            f.write(self._template_edit.toPlainText())
        logger.info("Exportação de npcs salva: %s (%s, %d npc(s))", path, fmt, len(self._npcs))

    def _on_file_dropped(self, path: str):
        if not self._uow:
            return
        suffix = path.rsplit(".", 1)[-1].lower() if "." in path else ""
        reader = {"json": self._read_json, "csv": self._read_csv, "xlsx": self._read_xlsx}.get(suffix)
        if reader is None:
            logger.warning("Formato de arquivo não suportado para import: %s", path)
            return
        try:
            data = reader(path)
        except Exception:
            logger.exception("Falha ao ler arquivo de importação: %s", path)
            return
        imported = self._import_npc_dicts(data)
        logger.info("Importados %d npcs de %s", imported, path)
        self._reload()
        self._close_tools_mode()

    def _export_xlsx(self):
        path, _selected = QFileDialog.getSaveFileName(
            self, "Exportar NPCs (Excel)", "npcs.xlsx", "Planilha Excel (*.xlsx)",
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation

        fieldnames = list(_NPC_TEMPLATE_FIELDS.keys())
        wb = Workbook()
        ws = wb.active
        ws.title = "NPCs"
        ws.append(fieldnames)
        for m in self._npcs:
            ws.append([m.get(k, _NPC_TEMPLATE_FIELDS[k]) for k in fieldnames])

        categories = sorted(self._all_categories, key=lambda c: c["name"])
        cats_ws = wb.create_sheet("Categorias")
        cats_ws.append(["id", "nome"])
        for c in categories:
            cats_ws.append([c["id"], c["name"]])

        if categories and "category" in fieldnames:
            col_letter = ws.cell(row=1, column=fieldnames.index("category") + 1).column_letter
            dv = DataValidation(type="list", formula1=f"=Categorias!$A$2:$A${len(categories) + 1}", allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}{len(self._npcs) + 1}")

        wb.save(path)
        logger.info("NPCs exportados para %s (%d linha(s))", path, len(self._npcs))

    def _export_xlsx_blank_template(self):
        path, _selected = QFileDialog.getSaveFileName(
            self, "Baixar Template de NPCs (Excel)", "npcs_template.xlsx", "Planilha Excel (*.xlsx)",
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation

        fieldnames = list(_NPC_TEMPLATE_FIELDS.keys())
        example = dict(_NPC_TEMPLATE_FIELDS)
        example["name"] = "Novo NPC"
        wb = Workbook()
        ws = wb.active
        ws.title = "NPCs"
        ws.append(fieldnames)
        ws.append([example.get(k, "") for k in fieldnames])

        categories = sorted(self._all_categories, key=lambda c: c["name"])
        cats_ws = wb.create_sheet("Categorias")
        cats_ws.append(["id", "nome"])
        for c in categories:
            cats_ws.append([c["id"], c["name"]])

        if categories and "category" in fieldnames:
            col_letter = ws.cell(row=1, column=fieldnames.index("category") + 1).column_letter
            dv = DataValidation(type="list", formula1=f"=Categorias!$A$2:$A${len(categories) + 1}", allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}1000")

        wb.save(path)
        logger.info("Template Excel de npcs baixado: %s", path)

    def _import_npc_dicts(self, data) -> int:
        if not isinstance(data, list):
            logger.warning("Arquivo de importação inválido: esperada uma lista de npcs.")
            return 0
        known_columns = self._known_npc_columns()
        if known_columns is None:
            logger.error("Import de npcs abortado: não foi possível confirmar as colunas válidas da tabela.")
            return 0
        imported = 0
        image_matches = 0
        for entry in data:
            if not isinstance(entry, dict) or not entry.get("name"):
                continue
            fields = {k: v for k, v in entry.items()
                      if k not in ("id", "created_at", "updated_at") and k in known_columns}
            if self._current_dir_id is not None:
                fields["category"] = self._current_dir_id
            npc_id = self._uow.npcs.create(**fields)
            imported += 1
            if not fields.get("image_path") and self._staged_image_files:
                key = _normalize_npc_name(fields.get("name", ""))
                source_path = self._staged_image_files.get(key)
                if source_path:
                    try:
                        rel_path = import_asset(self._project_dir, source_path, "assets/npcs", npc_id)
                        self._uow.npcs.update(npc_id, image_path=rel_path)
                        image_matches += 1
                    except Exception:
                        logger.exception("Falha ao atribuir imagem ao npc recém-criado %s (%s)", fields.get("name", ""), source_path)
        logger.info(
            "NPCs importados: %d (categoria=%s, %d imagem(ns) auto-atribuída(s))",
            imported, self._current_dir_id, image_matches,
        )
        return imported

    def _known_npc_columns(self) -> set[str] | None:
        if not self._uow:
            return None
        try:
            return set(self._uow.npcs.db.table_columns(self._uow.npcs.TABLE))
        except Exception:
            logger.exception("Falha ao introspectar colunas de npcs; import sem filtro de colunas.")
            return None

    def _read_json(self, path: str) -> list:
        return read_json(path)

    def _read_csv(self, path: str) -> list[dict]:
        return read_csv(path)

    @staticmethod
    def _normalize_blank_cells(row: dict) -> dict:
        return normalize_blank_cells(row)

    def _read_xlsx(self, path: str) -> list[dict]:
        return read_xlsx(path)
