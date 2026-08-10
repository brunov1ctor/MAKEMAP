"""Import/export building blocks shared by any panel that offers a
JSON/CSV/Excel paste-or-drop Importar flow plus folder-based image
matching — first written for Mobs (see mobs/panel_import_export_mixin.py),
extracted here so Items e Habilidades can reuse the same pieces instead of
re-implementing them. Everything here is a pure function or a dumb widget
with zero coupling to any specific panel's self.* state; the actual card-
building UI (which IS coupled to each panel's data) stays local to each
panel's own mixin.
"""

from __future__ import annotations

import csv
import io
import logging
import os
import re
import unicodedata

from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QDragEnterEvent, QDropEvent
from PySide6.QtWidgets import (
    QFrame, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFileDialog,
    QToolButton, QTextEdit, QWidget,
)

from src.styles.tokens import Colors

logger = logging.getLogger("MAKEMAP")


def normalize_name(text: str) -> str:
    """Case/accent/separator-insensitive comparison key — 'Salamandra_de_Fogo'
    and 'salamandra de fogo' normalize to the same value. Exact match only,
    no substring/fuzzy matching."""
    text = unicodedata.normalize("NFKD", text or "").encode("ascii", "ignore").decode()
    text = re.sub(r"[_\-\s]+", " ", text).strip().lower()
    return text


def index_files_by_stem(folder: str, extensions: tuple[str, ...]) -> dict[str, str]:
    """{normalized stem: full path} for every file directly in `folder`
    (no recursion) whose extension is in `extensions`."""
    try:
        entries = os.listdir(folder)
    except OSError:
        logger.exception("Falha ao listar pasta: %s", folder)
        return {}
    files_by_name: dict[str, str] = {}
    for entry in entries:
        path = os.path.join(folder, entry)
        stem, ext = os.path.splitext(entry)
        if ext.lower() in extensions and os.path.isfile(path):
            files_by_name[normalize_name(stem)] = path
    return files_by_name


def normalize_blank_cells(row: dict) -> dict:
    """CSV/Excel have no NULL — an unset value round-trips as a blank cell
    (empty string from csv, None from openpyxl) instead. Both forms are
    normalized to None here so a blank cell means "not set" on reimport,
    same as it did before export."""
    return {k: (None if v in ("", None) else v) for k, v in row.items()}


def read_json(path: str) -> list:
    import json
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def read_csv(path: str) -> list[dict]:
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        return [normalize_blank_cells(dict(row)) for row in csv.DictReader(f)]


def read_xlsx(path: str) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows, [])]
    return [
        normalize_blank_cells({header[i]: cell for i, cell in enumerate(row) if i < len(header)})
        for row in rows
    ]


def import_button_row(on_cancel, on_template, on_apply) -> tuple[QHBoxLayout, QPushButton]:
    """Cancelar/Template/Aplicar row shared by every paste-or-drop Importar
    card. Returns the Aplicar button too so a card can start it disabled
    until something is actually staged (e.g. a dropped file)."""
    row = QHBoxLayout()
    row.addStretch()
    cancel_btn = QPushButton("Cancelar")
    cancel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
    cancel_btn.setStyleSheet(f"""
        QPushButton {{ background: rgba(255,255,255,0.06); color: {Colors.TEXT_SECONDARY}; border: none;
            border-radius: 6px; padding: 5px 10px; font-size: 10px; }}
        QPushButton:hover {{ background: rgba(255,255,255,0.12); }}
    """)
    cancel_btn.clicked.connect(on_cancel)
    row.addWidget(cancel_btn)
    template_btn = QPushButton("Template")
    template_btn.setCursor(Qt.CursorShape.PointingHandCursor)
    template_btn.setStyleSheet(f"""
        QPushButton {{ background: rgba(255,255,255,0.06); color: {Colors.TEXT_PRIMARY};
            border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 6px; padding: 5px 10px; font-size: 10px; }}
        QPushButton:hover {{ background: rgba(255,255,255,0.12); border-color: {Colors.ACCENT}; }}
    """)
    template_btn.clicked.connect(on_template)
    row.addWidget(template_btn)
    apply_btn = QPushButton("Aplicar")
    apply_btn.setCursor(Qt.CursorShape.PointingHandCursor)
    apply_btn.setStyleSheet(f"""
        QPushButton {{ background: {Colors.ACCENT}; color: #08131F; border: none;
            border-radius: 6px; padding: 5px 12px; font-size: 10px; font-weight: bold; }}
        QPushButton:hover:enabled {{ background: {Colors.ACCENT_HOVER}; }}
        QPushButton:disabled {{ background: rgba(255,255,255,0.08); color: {Colors.TEXT_MUTED}; }}
    """)
    apply_btn.clicked.connect(on_apply)
    row.addWidget(apply_btn)
    return row, apply_btn


class DropZone(QFrame):
    """Big dashed-border area that accepts a dragged-in JSON/CSV/Excel file
    (or a plain click, opening the same multi-format file picker as a
    fallback for anyone who can't drag-and-drop)."""

    file_chosen = Signal(str)  # local file path, either dropped or picked

    _ACCEPTED_SUFFIXES = (".json", ".csv", ".xlsx")

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setStyleSheet(f"""
            QFrame {{ background: rgba(255,255,255,0.02); border: 2px dashed {Colors.BORDER_SUBTLE}; border-radius: 10px; }}
        """)
        lay = QVBoxLayout(self)
        lay.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.setSpacing(8)
        icon = QLabel("📥")
        icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        icon.setStyleSheet("font-size: 40px; background: transparent; border: none;")
        lay.addWidget(icon)
        hint = QLabel("Arraste um arquivo aqui\n(JSON, CSV ou Excel)")
        hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        hint.setStyleSheet(f"color: {Colors.TEXT_SECONDARY}; font-size: 11px; font-weight: bold; background: transparent; border: none;")
        lay.addWidget(hint)
        sub = QLabel("ou clique para escolher um arquivo")
        sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sub.setStyleSheet(f"color: {Colors.TEXT_MUTED}; font-size: 9px; background: transparent; border: none;")
        lay.addWidget(sub)

    def _first_accepted_path(self, mime_data) -> str | None:
        for url in mime_data.urls():
            path = url.toLocalFile()
            if path and path.lower().endswith(self._ACCEPTED_SUFFIXES):
                return path
        return None

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls() and self._first_accepted_path(event.mimeData()):
            event.acceptProposedAction()
            self.setStyleSheet(f"""
                QFrame {{ background: rgba(255,255,255,0.05); border: 2px dashed {Colors.ACCENT}; border-radius: 10px; }}
            """)
        else:
            event.ignore()

    def dragLeaveEvent(self, event):
        self.setStyleSheet(f"""
            QFrame {{ background: rgba(255,255,255,0.02); border: 2px dashed {Colors.BORDER_SUBTLE}; border-radius: 10px; }}
        """)
        super().dragLeaveEvent(event)

    def dropEvent(self, event: QDropEvent):
        path = self._first_accepted_path(event.mimeData())
        self.dragLeaveEvent(event)  # restore the idle style regardless of outcome
        if path:
            event.acceptProposedAction()
            self.file_chosen.emit(path)
        else:
            event.ignore()

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            path, _selected = QFileDialog.getOpenFileName(
                self, "Importar", "", "Todos os suportados (*.json *.csv *.xlsx);;JSON (*.json);;CSV (*.csv);;Excel (*.xlsx)",
            )
            if path:
                self.file_chosen.emit(path)
        super().mousePressEvent(event)


def build_tools_header(on_close, title_font_size: int = 13, close_font_size: int = 14) -> tuple[QHBoxLayout, QLabel, QToolButton]:
    """Title label + ✕ close button row shared by every entity panel's
    tools shell (_build_tools_panel) — identical everywhere except two
    font sizes, which vary between the "toggle" panels (Itens/Dungeons/
    Quests, 13/14) and the "single entity" panels (Mobs/NPCs, 12/12).
    Returns (head_row, title_label, close_button) — the caller adds
    head_row to its own `outer` layout and keeps the title_label reference
    (as self._tools_title_lbl) to update its text later."""
    head_row = QHBoxLayout()
    title_lbl = QLabel("")
    title_lbl.setStyleSheet(f"color: {Colors.TEXT_PRIMARY}; font-size: {title_font_size}px; font-weight: bold; background: transparent; border: none;")
    head_row.addWidget(title_lbl, 1)
    close_btn = QToolButton()
    close_btn.setText("✕")
    close_btn.setCursor(Qt.CursorShape.PointingHandCursor)
    close_btn.setToolTip("Fechar")
    close_btn.setStyleSheet(f"""
        QToolButton {{ border: none; background: transparent; color: {Colors.TEXT_MUTED}; font-size: {close_font_size}px; padding: 2px 6px; }}
        QToolButton:hover {{ color: {Colors.TEXT_PRIMARY}; }}
        QToolTip {{
            background-color: {Colors.BG_ELEVATED};
            color: {Colors.TEXT_PRIMARY};
            border: 1px solid {Colors.BORDER};
            border-radius: 8px;
            padding: 6px 10px;
            font-size: 11px;
        }}
    """)
    close_btn.clicked.connect(on_close)
    head_row.addWidget(close_btn)
    return head_row, title_lbl, close_btn


def build_export_view_page(on_save) -> tuple[QWidget, QLabel, QTextEdit]:
    """Page 1 of the tools stack (see _build_tools_panel) — the read-only
    Exportar view: a hint label, a read-only monospace text box showing
    the export content, and a "Salvar Arquivo" button. Byte-identical
    across every entity panel. Returns (page, hint_label, text_edit) — the
    caller adds `page` to its own QStackedWidget and keeps the two live
    widgets (as self._template_hint_lbl/self._template_edit) to fill in
    on export."""
    page = QWidget()
    lay = QVBoxLayout(page)
    lay.setContentsMargins(0, 0, 0, 0)
    lay.setSpacing(6)

    hint_lbl = QLabel("")
    hint_lbl.setWordWrap(True)
    hint_lbl.setStyleSheet(f"color: {Colors.TEXT_MUTED}; font-size: 9px; background: transparent; border: none;")
    lay.addWidget(hint_lbl)

    text_edit = QTextEdit()
    text_edit.setReadOnly(True)
    text_edit.setStyleSheet(f"""
        QTextEdit {{ color: {Colors.TEXT_PRIMARY}; font-size: 10px; font-family: Consolas, monospace;
            background: rgba(0,0,0,0.2); border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 4px; padding: 6px; }}
    """)
    lay.addWidget(text_edit, 1)

    btn_row = QHBoxLayout()
    btn_row.addStretch()
    save_btn = QPushButton("💾 Salvar Arquivo")
    save_btn.setCursor(Qt.CursorShape.PointingHandCursor)
    save_btn.setStyleSheet(f"""
        QPushButton {{ background: {Colors.ACCENT}; color: #08131F; border: none;
            border-radius: 6px; padding: 6px 14px; font-size: 10px; font-weight: bold; }}
        QPushButton:hover {{ background: {Colors.ACCENT_HOVER}; }}
    """)
    save_btn.clicked.connect(on_save)
    btn_row.addWidget(save_btn)
    lay.addLayout(btn_row)

    return page, hint_lbl, text_edit
