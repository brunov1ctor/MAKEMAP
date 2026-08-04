"""ImportExportMixin — the Importar (JSON/CSV paste cards + Excel drop
card) / Exportar (read-only JSON/CSV view + direct-to-file Excel) tools
panel that takes over the right column while active. Mixed into MobsPanel
(see panel.py) — operates on self.* attributes MobsPanel owns; not meant to
be instantiated on its own.
"""

from __future__ import annotations

import logging

from PySide6.QtWidgets import (
    QFrame, QVBoxLayout, QHBoxLayout, QLabel, QToolButton, QTextEdit,
    QPushButton, QWidget, QStackedWidget, QScrollArea, QFileDialog,
)
from PySide6.QtCore import Qt

from src.styles.tokens import Colors
from src.layouts.panels.mobs.mob_edit_panel import MobEditPanel
from src.layouts.panels.mobs.panel_widgets import _DropZone
from src.layouts.panels.mobs.panel_helpers import (
    _MOB_TEMPLATE_FIELDS, _TEMPLATE_FIELD_DOCS, _parse_mobs_json, _normalize_mob_name,
)
from src.layouts.panels.shared.import_export_helpers import (
    index_files_by_stem, normalize_blank_cells, read_json, read_csv, read_xlsx,
    import_button_row,
)
from src.services.project_assets import import_asset

_IMAGE_EXTS = (".png", ".jpg", ".jpeg", ".webp", ".bmp")
_ASSET_EXTS = (".fbx", ".obj", ".gltf", ".glb", ".png", ".jpg", ".jpeg", ".webp")

logger = logging.getLogger("MAKEMAP")


class ImportExportMixin:
    """The right column's alternate page (see self._right_stack) — a
    header (title + ✕ close) above a small stack of its own: page 0 is
    Importar — separate JSON / CSV (paste text) / Excel (drop file) cards,
    each with its own Cancelar/Template/Aplicar row — page 1 is a
    read-only JSON/CSV export view of the user's current mobs (Exportar)."""

    def _build_tools_panel(self) -> QWidget:
        panel = QFrame()
        panel.setStyleSheet(f"QFrame {{ background: rgba(255,255,255,0.03); border-radius: 8px; }}")
        outer = QVBoxLayout(panel)
        outer.setContentsMargins(10, 10, 10, 10)
        outer.setSpacing(8)

        head_row = QHBoxLayout()
        self._tools_title_lbl = QLabel("")
        self._tools_title_lbl.setStyleSheet(f"color: {Colors.TEXT_PRIMARY}; font-size: 12px; font-weight: bold; background: transparent; border: none;")
        head_row.addWidget(self._tools_title_lbl, 1)
        tools_close_btn = QToolButton()
        tools_close_btn.setText("✕")
        tools_close_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        tools_close_btn.setToolTip("Fechar")
        tools_close_btn.setStyleSheet(f"""
            QToolButton {{ border: none; background: transparent; color: {Colors.TEXT_MUTED}; font-size: 12px; padding: 2px 6px; }}
            QToolButton:hover {{ color: {Colors.TEXT_PRIMARY}; }}
            QToolTip {{
                background-color: {Colors.BG_ELEVATED};
                color: {Colors.TEXT_PRIMARY};
                border: 1px solid {Colors.BORDER};
                border-radius: 8px;
                padding: 6px 10px;
                font-size: 11px;
            }}
        """)
        tools_close_btn.clicked.connect(self._close_tools_mode)
        head_row.addWidget(tools_close_btn)
        outer.addLayout(head_row)

        self._tools_stack = QStackedWidget()

        # ── Page 0: Importar — JSON / CSV / Excel cards, stacked ──
        import_content = QWidget()
        import_lay = QVBoxLayout(import_content)
        import_lay.setContentsMargins(0, 0, 0, 0)
        import_lay.setSpacing(8)
        import_lay.addWidget(self._build_text_import_card(
            "JSON", "Cole uma lista JSON para criar vários mobs de uma vez.", "json"))
        import_lay.addWidget(self._build_text_import_card(
            "CSV", "Cole uma lista CSV para criar vários mobs de uma vez.", "csv"))
        import_lay.addWidget(self._build_excel_import_card())
        import_lay.addWidget(self._build_image_folder_import_card())
        import_lay.addWidget(self._build_asset_folder_import_card())
        import_lay.addStretch()
        import_scroll = QScrollArea()
        import_scroll.setWidgetResizable(True)
        import_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        import_scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }"
                                     "QScrollArea > QWidget > QWidget { background: transparent; }")
        import_scroll.setWidget(import_content)
        self._tools_stack.addWidget(import_scroll)

        # ── Page 1: read-only JSON/CSV export view ──
        template_page = QWidget()
        template_lay = QVBoxLayout(template_page)
        template_lay.setContentsMargins(0, 0, 0, 0)
        template_lay.setSpacing(6)

        self._template_hint_lbl = QLabel("")
        self._template_hint_lbl.setWordWrap(True)
        self._template_hint_lbl.setStyleSheet(f"color: {Colors.TEXT_MUTED}; font-size: 9px; background: transparent; border: none;")
        template_lay.addWidget(self._template_hint_lbl)

        # Read-only — this page is a live view of the user's own mobs (see
        # _build_mob_export), not an editable form you "apply": there's
        # nothing here to submit, just content to copy or save to a file.
        self._template_edit = QTextEdit()
        self._template_edit.setReadOnly(True)
        self._template_edit.setStyleSheet(f"""
            QTextEdit {{ color: {Colors.TEXT_PRIMARY}; font-size: 10px; font-family: Consolas, monospace;
                background: rgba(0,0,0,0.2); border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 4px; padding: 6px; }}
        """)
        template_lay.addWidget(self._template_edit, 1)

        template_btn_row = QHBoxLayout()
        template_btn_row.addStretch()
        template_save_btn = QPushButton("💾 Salvar Arquivo")
        template_save_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        template_save_btn.setStyleSheet(f"""
            QPushButton {{ background: {Colors.ACCENT}; color: #08131F; border: none;
                border-radius: 6px; padding: 6px 14px; font-size: 10px; font-weight: bold; }}
            QPushButton:hover {{ background: {Colors.ACCENT_HOVER}; }}
        """)
        template_save_btn.clicked.connect(self._on_save_export_file)
        template_btn_row.addWidget(template_save_btn)
        template_lay.addLayout(template_btn_row)

        self._tools_stack.addWidget(template_page)
        outer.addWidget(self._tools_stack, 1)

        panel.setMinimumWidth(MobEditPanel.PANEL_WIDTH)
        return panel

    # ─── Importar cards (JSON/CSV paste, Excel drop) ───

    def _import_button_row(self, on_cancel, on_template, on_apply) -> tuple[QHBoxLayout, QPushButton]:
        """Cancelar/Template/Aplicar row shared by all Importar cards —
        see shared.import_export_helpers.import_button_row."""
        return import_button_row(on_cancel, on_template, on_apply)

    def _json_import_template(self) -> str:
        """The immutable base template the JSON card's "Template" button
        always resets back to — a single documented example mob, not the
        user's real data (Aplicar always CREATES new rows here, so
        prefilling real data would risk silently duplicating it)."""
        import json
        valid_categories = ", ".join(
            f"{c['id']} ({c['name']})" for c in sorted(self._all_categories, key=lambda c: c["name"])
        ) or "nenhuma ainda — crie uma no explorador à esquerda"
        doc_lines = "\n".join(f"// {key}: {doc}" for key, doc in _TEMPLATE_FIELD_DOCS)
        cat_line = f"// Categorias disponíveis: {valid_categories}"
        example = dict(_MOB_TEMPLATE_FIELDS)
        example["name"] = "Novo Mob"
        row_json = json.dumps(example, ensure_ascii=False, indent=2)
        row_indented = "\n".join(f"  {line}" for line in row_json.splitlines())
        return f"{doc_lines}\n{cat_line}\n[\n{row_indented}\n]"

    def _csv_import_template(self) -> str:
        """CSV equivalent of _json_import_template — header row + one
        example mob row."""
        import csv
        import io
        example = dict(_MOB_TEMPLATE_FIELDS)
        example["name"] = "Novo Mob"
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=list(_MOB_TEMPLATE_FIELDS.keys()))
        writer.writeheader()
        writer.writerow(example)
        return buf.getvalue()

    def _build_text_import_card(self, title: str, hint: str, fmt: str) -> QFrame:
        """One JSON or CSV paste-and-create card — Template resets the
        text back to the immutable base example, Aplicar parses whatever
        is currently in the box (edited or not) and creates a mob per
        entry."""
        card = QFrame()
        card.setStyleSheet(f"QFrame {{ background: rgba(0,0,0,0.15); border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 8px; }}")
        lay = QVBoxLayout(card)
        lay.setContentsMargins(10, 8, 10, 8)
        lay.setSpacing(5)

        title_lbl = QLabel(title)
        title_lbl.setStyleSheet(f"color: {Colors.TEXT_PRIMARY}; font-size: 11px; font-weight: bold; background: transparent; border: none;")
        lay.addWidget(title_lbl)
        hint_lbl = QLabel(hint)
        hint_lbl.setWordWrap(True)
        hint_lbl.setStyleSheet(f"color: {Colors.TEXT_MUTED}; font-size: 9px; background: transparent; border: none;")
        lay.addWidget(hint_lbl)

        edit = QTextEdit()
        edit.setFixedHeight(120)
        edit.setStyleSheet(f"""
            QTextEdit {{ color: {Colors.TEXT_PRIMARY}; font-size: 9px; font-family: Consolas, monospace;
                background: rgba(0,0,0,0.25); border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 4px; padding: 5px; }}
        """)
        lay.addWidget(edit)

        error_lbl = QLabel("")
        error_lbl.setWordWrap(True)
        error_lbl.setStyleSheet(f"color: {Colors.ERROR}; font-size: 9px; background: transparent; border: none;")
        error_lbl.hide()
        lay.addWidget(error_lbl)

        def reset_to_template():
            edit.setPlainText(self._json_import_template() if fmt == "json" else self._csv_import_template())
            error_lbl.hide()

        def do_apply():
            text = edit.toPlainText()
            try:
                if fmt == "json":
                    data = _parse_mobs_json(text)
                else:
                    import csv
                    import io
                    data = [self._normalize_blank_cells(dict(row)) for row in csv.DictReader(io.StringIO(text))]
            except ValueError as exc:
                error_lbl.setText(str(exc))
                error_lbl.show()
                return
            except Exception:
                logger.exception("Falha ao interpretar import de mobs (%s).", fmt)
                error_lbl.setText("Não foi possível interpretar o conteúdo — confira o formato.")
                error_lbl.show()
                return
            imported = self._import_mob_dicts(data)
            if imported == 0:
                error_lbl.setText('Nenhum mob válido encontrado (o campo "name" é obrigatório).')
                error_lbl.show()
                return
            logger.info("Criados %d mob(s) via Importar (%s)", imported, fmt)
            self._reload()
            reset_to_template()
            self._close_tools_mode()

        row, _apply_btn = self._import_button_row(self._close_tools_mode, reset_to_template, do_apply)
        lay.addLayout(row)
        reset_to_template()
        return card

    def _build_excel_import_card(self) -> QFrame:
        """The Excel card — same drop/click zone as before, but dropping
        or picking a file now just STAGES it (shows the filename, enables
        Aplicar) instead of importing immediately, matching the JSON/CSV
        cards' explicit Cancelar/Template/Aplicar flow. Template downloads
        a blank starting .xlsx (headers + one example row + the
        Categorias validation sheet) instead of resetting an editable
        text box, since there's nothing to paste here."""
        card = QFrame()
        card.setStyleSheet(f"QFrame {{ background: rgba(0,0,0,0.15); border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 8px; }}")
        lay = QVBoxLayout(card)
        lay.setContentsMargins(10, 8, 10, 8)
        lay.setSpacing(5)

        title_lbl = QLabel("Excel")
        title_lbl.setStyleSheet(f"color: {Colors.TEXT_PRIMARY}; font-size: 11px; font-weight: bold; background: transparent; border: none;")
        lay.addWidget(title_lbl)

        drop_zone = _DropZone()
        lay.addWidget(drop_zone)

        staged_lbl = QLabel("Nenhum arquivo selecionado.")
        staged_lbl.setWordWrap(True)
        staged_lbl.setStyleSheet(f"color: {Colors.TEXT_MUTED}; font-size: 9px; background: transparent; border: none;")
        lay.addWidget(staged_lbl)

        state = {"path": ""}

        def on_file_staged(path: str):
            state["path"] = path
            import os
            staged_lbl.setText(f"Selecionado: {os.path.basename(path)}")
            apply_btn.setEnabled(True)

        drop_zone.file_chosen.connect(on_file_staged)

        def clear_staged():
            state["path"] = ""
            staged_lbl.setText("Nenhum arquivo selecionado.")
            apply_btn.setEnabled(False)

        def do_cancel():
            clear_staged()
            self._close_tools_mode()

        def do_apply():
            if not state["path"]:
                return
            self._on_file_dropped(state["path"])
            clear_staged()

        row, apply_btn = self._import_button_row(do_cancel, self._export_xlsx_blank_template, do_apply)
        apply_btn.setEnabled(False)
        lay.addLayout(row)
        return card

    def _build_image_folder_import_card(self) -> QFrame:
        """Match a folder of portrait images against existing mobs by
        normalized filename == normalized mob name, then attach the
        matched image to each mob — same stage-then-Aplicar flow as the
        Excel card (folder pick just builds a preview; nothing is copied/
        saved until Aplicar), except there's no Template button (nothing
        to hand-fill here)."""
        card = QFrame()
        card.setStyleSheet(f"QFrame {{ background: rgba(0,0,0,0.15); border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 8px; }}")
        lay = QVBoxLayout(card)
        lay.setContentsMargins(10, 8, 10, 8)
        lay.setSpacing(5)

        title_lbl = QLabel("Imagens (pasta)")
        title_lbl.setStyleSheet(f"color: {Colors.TEXT_PRIMARY}; font-size: 11px; font-weight: bold; background: transparent; border: none;")
        lay.addWidget(title_lbl)

        hint_lbl = QLabel(
            "Escolha uma pasta com imagens nomeadas como os mobs (ex.: "
            "\"Salamandra de Fogo.png\") — cada arquivo cujo nome bater com "
            "um mob existente recebe essa imagem como retrato."
        )
        hint_lbl.setWordWrap(True)
        hint_lbl.setStyleSheet(f"color: {Colors.TEXT_MUTED}; font-size: 9px; background: transparent; border: none;")
        lay.addWidget(hint_lbl)

        pick_btn = QPushButton("📁 Selecionar Pasta")
        pick_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        pick_btn.setStyleSheet(f"""
            QPushButton {{ background: rgba(255,255,255,0.06); color: {Colors.TEXT_PRIMARY};
                border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 6px; padding: 5px 10px; font-size: 10px; }}
            QPushButton:hover {{ background: rgba(255,255,255,0.12); border-color: {Colors.ACCENT}; }}
        """)
        lay.addWidget(pick_btn)

        preview_edit = QTextEdit()
        preview_edit.setReadOnly(True)
        preview_edit.setFixedHeight(120)
        preview_edit.setStyleSheet(f"""
            QTextEdit {{ color: {Colors.TEXT_PRIMARY}; font-size: 9px; font-family: Consolas, monospace;
                background: rgba(0,0,0,0.2); border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 4px; padding: 6px; }}
        """)
        preview_edit.setPlainText("Nenhuma pasta selecionada.")
        lay.addWidget(preview_edit)

        state = {"matches": []}  # list of (mob_dict, source_file_path)

        def refresh_preview():
            matches, unmatched = self._match_mobs_against_staged_folder()
            state["matches"] = matches
            preview_edit.setPlainText(self._format_image_match_preview(matches, unmatched))
            apply_btn.setEnabled(bool(matches))

        def on_pick_folder():
            folder = QFileDialog.getExistingDirectory(self, "Selecione a pasta com as imagens dos mobs")
            if not folder:
                return
            # Kept on self (not just this closure's `state`) so a
            # JSON/CSV/Excel import run afterward — without reopening this
            # card — can auto-match each newly created mob against the
            # same folder (see _import_mob_dicts).
            self._staged_image_folder = folder
            self._staged_image_files = self._index_image_folder(folder)
            refresh_preview()

        pick_btn.clicked.connect(on_pick_folder)

        def clear_staged():
            state["matches"] = []
            preview_edit.setPlainText("Nenhuma pasta selecionada.")
            apply_btn.setEnabled(False)

        def do_cancel():
            # Cancelar only clears THIS card's preview/Aplicar state — the
            # staged folder itself stays put, since the whole point is
            # letting the user leave it selected and go create mobs via
            # JSON/CSV/Excel instead, without losing the pick.
            state["matches"] = []
            self._close_tools_mode()

        def do_apply():
            if not state["matches"]:
                return
            self._apply_image_matches(state["matches"])
            clear_staged()

        row = QHBoxLayout()
        row.addStretch()
        cancel_btn = QPushButton("Cancelar")
        cancel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        cancel_btn.setStyleSheet(f"""
            QPushButton {{ background: rgba(255,255,255,0.06); color: {Colors.TEXT_SECONDARY}; border: none;
                border-radius: 6px; padding: 5px 10px; font-size: 10px; }}
            QPushButton:hover {{ background: rgba(255,255,255,0.12); }}
        """)
        cancel_btn.clicked.connect(do_cancel)
        row.addWidget(cancel_btn)
        apply_btn = QPushButton("Aplicar")
        apply_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        apply_btn.setStyleSheet(f"""
            QPushButton {{ background: {Colors.ACCENT}; color: #08131F; border: none;
                border-radius: 6px; padding: 5px 12px; font-size: 10px; font-weight: bold; }}
            QPushButton:hover:enabled {{ background: {Colors.ACCENT_HOVER}; }}
            QPushButton:disabled {{ background: rgba(255,255,255,0.08); color: {Colors.TEXT_MUTED}; }}
        """)
        apply_btn.setEnabled(False)
        apply_btn.clicked.connect(do_apply)
        row.addWidget(apply_btn)
        lay.addLayout(row)
        return card

    @staticmethod
    def _index_image_folder(folder: str) -> dict[str, str]:
        """{normalized stem: full path} for every image file directly in
        `folder` (no recursion) — the shared index both the preview
        (matched against CURRENT mobs) and _import_mob_dicts (matched
        against mobs created AFTER the folder was staged) read from."""
        return index_files_by_stem(folder, _IMAGE_EXTS)

    def _match_mobs_against_staged_folder(self) -> tuple[list, list[str]]:
        """Normalized-exact match only (see _normalize_mob_name) — no
        substring/fuzzy matching, so a near-miss shows up as unmatched
        instead of silently attaching the wrong image. Matches self._mobs
        (as they are RIGHT NOW) against self._staged_image_files."""
        import os
        files_by_name = self._staged_image_files
        matches = []
        matched_keys = set()
        for m in self._mobs:
            key = _normalize_mob_name(m.get("name", ""))
            if key and key in files_by_name:
                matches.append((m, files_by_name[key]))
                matched_keys.add(key)

        unmatched = [
            os.path.basename(path) for key, path in files_by_name.items()
            if key not in matched_keys
        ]
        return matches, sorted(unmatched)

    @staticmethod
    def _format_image_match_preview(matches: list, unmatched: list[str]) -> str:
        import os
        lines = [f"{len(matches) + len(unmatched)} imagem(ns) encontradas, {len(matches)} mob(s) correspondente(s):"]
        for mob, path in matches:
            suffix = " (substituirá imagem atual)" if mob.get("image_path") else ""
            lines.append(f"  OK {mob.get('name', '')}  <-  {os.path.basename(path)}{suffix}")
        if unmatched:
            lines.append(f"{len(unmatched)} arquivo(s) sem mob correspondente:")
            for name in unmatched:
                lines.append(f"  - {name}")
        return "\n".join(lines)

    def _apply_image_matches(self, matches: list):
        if not self._uow:
            return
        applied = 0
        for mob, source_path in matches:
            mob_id = mob.get("id")
            if not mob_id:
                continue
            try:
                rel_path = import_asset(self._project_dir, source_path, "assets/mobs", mob_id)
                self._uow.mobs.update(mob_id, image_path=rel_path)
                applied += 1
            except Exception:
                logger.exception("Falha ao atribuir imagem ao mob %s (%s)", mob.get("name", ""), source_path)
        logger.info("Imagens atribuídas via pasta: %d mob(s)", applied)
        self._reload()
        self._close_tools_mode()

    # ─── Assets (pasta) — bulk mob_assets import, mirrors the image-folder
    # card above (pick → preview → Aplicar), but matches by SUBSTRING
    # instead of exact stem, and one mob can pick up several files at
    # once (a model + its textures), not just one. ───

    def _build_asset_folder_import_card(self) -> QFrame:
        """Match a folder of asset files (3D models/textures) against
        existing mobs and attach each match as a new mob_assets row —
        same stage-then-Aplicar flow as "Imagens (pasta)" above."""
        card = QFrame()
        card.setStyleSheet(f"QFrame {{ background: rgba(0,0,0,0.15); border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 8px; }}")
        lay = QVBoxLayout(card)
        lay.setContentsMargins(10, 8, 10, 8)
        lay.setSpacing(5)

        title_lbl = QLabel("Assets (pasta)")
        title_lbl.setStyleSheet(f"color: {Colors.TEXT_PRIMARY}; font-size: 11px; font-weight: bold; background: transparent; border: none;")
        lay.addWidget(title_lbl)

        hint_lbl = QLabel(
            "Escolha uma pasta com modelos/texturas cujo nome CONTENHA o "
            "nome de um mob (ex.: \"tex_goblin_01.jpg\", \"model_goblin.fbx\") "
            "— cada arquivo é adicionado como um novo asset do mob "
            "correspondente; um mob pode receber vários arquivos de uma vez."
        )
        hint_lbl.setWordWrap(True)
        hint_lbl.setStyleSheet(f"color: {Colors.TEXT_MUTED}; font-size: 9px; background: transparent; border: none;")
        lay.addWidget(hint_lbl)

        pick_btn = QPushButton("📁 Selecionar Pasta")
        pick_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        pick_btn.setStyleSheet(f"""
            QPushButton {{ background: rgba(255,255,255,0.06); color: {Colors.TEXT_PRIMARY};
                border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 6px; padding: 5px 10px; font-size: 10px; }}
            QPushButton:hover {{ background: rgba(255,255,255,0.12); border-color: {Colors.ACCENT}; }}
        """)
        lay.addWidget(pick_btn)

        preview_edit = QTextEdit()
        preview_edit.setReadOnly(True)
        preview_edit.setFixedHeight(120)
        preview_edit.setStyleSheet(f"""
            QTextEdit {{ color: {Colors.TEXT_PRIMARY}; font-size: 9px; font-family: Consolas, monospace;
                background: rgba(0,0,0,0.2); border: 1px solid {Colors.BORDER_SUBTLE}; border-radius: 4px; padding: 6px; }}
        """)
        preview_edit.setPlainText("Nenhuma pasta selecionada.")
        lay.addWidget(preview_edit)

        state = {"matches": []}  # list of (mob_dict, [source_file_paths])

        def refresh_preview():
            matches, unmatched = self._match_mobs_against_staged_asset_folder()
            state["matches"] = matches
            preview_edit.setPlainText(self._format_asset_match_preview(matches, unmatched))
            apply_btn.setEnabled(bool(matches))

        def on_pick_folder():
            folder = QFileDialog.getExistingDirectory(self, "Selecione a pasta com os assets dos mobs")
            if not folder:
                return
            self._staged_asset_folder = folder
            self._staged_asset_files = self._index_asset_folder(folder)
            refresh_preview()

        pick_btn.clicked.connect(on_pick_folder)

        def clear_staged():
            state["matches"] = []
            preview_edit.setPlainText("Nenhuma pasta selecionada.")
            apply_btn.setEnabled(False)

        def do_cancel():
            # Same reasoning as the image-folder card's Cancelar: only
            # this card's own preview/Aplicar state is cleared — the
            # staged folder itself stays put in case the user wants to
            # reopen this card later without reselecting it.
            state["matches"] = []
            self._close_tools_mode()

        def do_apply():
            if not state["matches"]:
                return
            self._apply_asset_matches(state["matches"])
            clear_staged()

        row = QHBoxLayout()
        row.addStretch()
        cancel_btn = QPushButton("Cancelar")
        cancel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        cancel_btn.setStyleSheet(f"""
            QPushButton {{ background: rgba(255,255,255,0.06); color: {Colors.TEXT_SECONDARY}; border: none;
                border-radius: 6px; padding: 5px 10px; font-size: 10px; }}
            QPushButton:hover {{ background: rgba(255,255,255,0.12); }}
        """)
        cancel_btn.clicked.connect(do_cancel)
        row.addWidget(cancel_btn)
        apply_btn = QPushButton("Aplicar")
        apply_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        apply_btn.setStyleSheet(f"""
            QPushButton {{ background: {Colors.ACCENT}; color: #08131F; border: none;
                border-radius: 6px; padding: 5px 12px; font-size: 10px; font-weight: bold; }}
            QPushButton:hover:enabled {{ background: {Colors.ACCENT_HOVER}; }}
            QPushButton:disabled {{ background: rgba(255,255,255,0.08); color: {Colors.TEXT_MUTED}; }}
        """)
        apply_btn.setEnabled(False)
        apply_btn.clicked.connect(do_apply)
        row.addWidget(apply_btn)
        lay.addLayout(row)
        return card

    @staticmethod
    def _index_asset_folder(folder: str) -> dict[str, list[str]]:
        """{normalized stem: [full paths]} for every asset file (3D model
        or texture) directly in `folder` (no recursion) — a list per key
        (not a single path) since two files can normalize to the same
        stem with different extensions (e.g. "goblin.fbx" + "goblin.png"),
        which the image-folder index's 1-path-per-key shape would silently
        drop one of."""
        import os
        try:
            entries = os.listdir(folder)
        except OSError:
            logger.exception("Falha ao listar pasta de assets: %s", folder)
            return {}
        files_by_name: dict[str, list[str]] = {}
        for entry in entries:
            path = os.path.join(folder, entry)
            stem, ext = os.path.splitext(entry)
            if ext.lower() in _ASSET_EXTS and os.path.isfile(path):
                files_by_name.setdefault(_normalize_mob_name(stem), []).append(path)
        return files_by_name

    def _match_mobs_against_staged_asset_folder(self) -> tuple[list, list[str]]:
        """SUBSTRING match (normalized mob name found inside the
        normalized filename) — unlike portraits, asset filenames almost
        always decorate the mob name with prefixes/suffixes ("tex_goblin_
        01.jpg", "model_goblin.fbx"), so requiring an exact stem match
        (like _match_mobs_against_staged_folder does for images) would
        essentially never hit. One mob can match several files (a model
        plus multiple textures); a file that doesn't contain ANY mob's
        name stays unmatched. Trade-off: a mob whose name is a substring
        of another mob's (rare) could over-match — acceptable given the
        preview always shows exactly what matched before Aplicar."""
        import os
        files_by_name = self._staged_asset_files
        matches: list[tuple[dict, list[str]]] = []
        matched_paths: set[str] = set()
        for m in self._mobs:
            key = _normalize_mob_name(m.get("name", ""))
            if not key:
                continue
            mob_files = sorted(
                path for norm_stem, paths in files_by_name.items() if key in norm_stem for path in paths
            )
            if mob_files:
                matches.append((m, mob_files))
                matched_paths.update(mob_files)

        unmatched = [
            os.path.basename(path) for paths in files_by_name.values() for path in paths
            if path not in matched_paths
        ]
        return matches, sorted(unmatched)

    @staticmethod
    def _format_asset_match_preview(matches: list, unmatched: list[str]) -> str:
        import os
        total_files = sum(len(paths) for _m, paths in matches) + len(unmatched)
        lines = [f"{total_files} arquivo(s) encontrados, {len(matches)} mob(s) correspondente(s):"]
        for mob, paths in matches:
            lines.append(f"  OK {mob.get('name', '')}  ({len(paths)} arquivo(s)):")
            for path in paths:
                lines.append(f"      - {os.path.basename(path)}")
        if unmatched:
            lines.append(f"{len(unmatched)} arquivo(s) sem mob correspondente:")
            for name in unmatched:
                lines.append(f"  - {name}")
        return "\n".join(lines)

    def _apply_asset_matches(self, matches: list):
        if not self._uow:
            return
        import os
        import uuid
        applied = 0
        affected_mob_ids = set()
        for mob, paths in matches:
            mob_id = mob.get("id")
            if not mob_id:
                continue
            for path in paths:
                try:
                    asset_id = str(uuid.uuid4())
                    ext = os.path.splitext(path)[1].lower()
                    asset_type = {
                        ".fbx": "Modelo 3D", ".obj": "Modelo 3D", ".gltf": "Modelo 3D", ".glb": "Modelo 3D",
                        ".png": "Imagem", ".jpg": "Imagem", ".jpeg": "Imagem", ".webp": "Imagem",
                    }.get(ext, "Arquivo")
                    try:
                        size = os.path.getsize(path)
                    except OSError:
                        size = 0
                    rel_path = import_asset(self._project_dir, path, "assets/mob_assets", asset_id)
                    self._uow.mob_assets.create(
                        id=asset_id, mob_id=mob_id, name=os.path.basename(path),
                        asset_type=asset_type, file_path=rel_path, file_size=size, rarity="common",
                    )
                    applied += 1
                    affected_mob_ids.add(mob_id)
                except Exception:
                    logger.exception("Falha ao importar asset para o mob %s (%s)", mob.get("name", ""), path)
        logger.info("Assets importados via pasta: %d arquivo(s) em %d mob(s)", applied, len(affected_mob_ids))
        if self._selected_id in affected_mob_ids:
            self._refresh_assets_display(self._selected_id)
        self._reload()
        self._close_tools_mode()

    def _toggle_import_mode(self):
        if self._tools_mode == "import":
            self._close_tools_mode()
            return
        self._tools_title_lbl.setText("Importar Mobs")
        self._tools_stack.setCurrentIndex(0)
        self._right_stack.setCurrentIndex(1)
        self._tools_mode = "import"

    def _on_export_choice(self, fmt: str):
        if fmt == "xlsx":
            # Not text-editable in-panel (it's a binary spreadsheet), so
            # this writes the FILE directly instead of taking over the
            # right panel — see _export_xlsx.
            self._export_xlsx()
            return
        mode = f"export_{fmt}"
        if self._tools_mode == mode:
            self._close_tools_mode()
            return
        self._template_fmt = fmt
        self._template_edit.setPlainText(self._build_mob_export(fmt))
        self._template_hint_lbl.setText(
            f"{len(self._mobs)} mob(s) — clique em Salvar Arquivo para exportar, ou copie o texto abaixo."
        )
        self._tools_title_lbl.setText("Exportar como JSON" if fmt == "json" else "Exportar como CSV")
        self._tools_stack.setCurrentIndex(1)
        self._right_stack.setCurrentIndex(1)
        self._tools_mode = mode

    def _close_tools_mode(self):
        self._right_stack.setCurrentIndex(0)
        self._tools_mode = None

    def _build_mob_export(self, fmt: str) -> str:
        """The user's actual current mobs (self._mobs — same data backing
        the grid), not a blank example row — this page used to always
        show one fixed placeholder mob no matter what had been created,
        since Aplicar (now removed) treated it as a create-from-template
        form rather than a real export."""
        fieldnames = list(_MOB_TEMPLATE_FIELDS.keys())
        rows = [{k: m.get(k, _MOB_TEMPLATE_FIELDS[k]) for k in fieldnames} for m in self._mobs]
        if fmt == "json":
            import json
            return json.dumps(rows, ensure_ascii=False, indent=2)
        # csv
        import csv
        import io
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
        return buf.getvalue()

    def _on_save_export_file(self):
        fmt = self._template_fmt
        ext = "json" if fmt == "json" else "csv"
        file_filter = "JSON (*.json)" if fmt == "json" else "CSV (*.csv)"
        path, _selected = QFileDialog.getSaveFileName(self, "Salvar Exportação de Mobs", f"mobs.{ext}", file_filter)
        if not path:
            return
        if not path.lower().endswith(f".{ext}"):
            path += f".{ext}"
        with open(path, "w", encoding="utf-8", newline="") as f:
            f.write(self._template_edit.toPlainText())
        logger.info("Exportação de mobs salva: %s (%s, %d mob(s))", path, fmt, len(self._mobs))

    def _on_file_dropped(self, path: str):
        if not self._uow:
            return
        suffix = path.rsplit(".", 1)[-1].lower() if "." in path else ""
        reader = {"json": self._read_json, "csv": self._read_csv, "xlsx": self._read_xlsx}.get(suffix)
        if reader is None:
            logger.warning("Formato de arquivo não suportado para import: %s", path)
            return
        try:
            data = reader(path)
        except Exception:
            logger.exception("Falha ao ler arquivo de importação: %s", path)
            return
        imported = self._import_mob_dicts(data)
        logger.info("Importados %d mobs de %s", imported, path)
        self._reload()
        self._close_tools_mode()

    def _export_xlsx(self):
        """The user's actual current mobs (self._mobs) — one row per mob,
        not a blank skeleton with a single example row — plus a
        Categorias reference sheet with a dropdown data-validation on the
        "category" column restricted to real category ids, so hand-
        editing the file afterward can't produce a typo'd/nonexistent
        category."""
        path, _selected = QFileDialog.getSaveFileName(
            self, "Exportar Mobs (Excel)", "mobs.xlsx", "Planilha Excel (*.xlsx)",
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation

        fieldnames = list(_MOB_TEMPLATE_FIELDS.keys())
        wb = Workbook()
        ws = wb.active
        ws.title = "Mobs"
        ws.append(fieldnames)
        for m in self._mobs:
            ws.append([m.get(k, _MOB_TEMPLATE_FIELDS[k]) for k in fieldnames])

        categories = sorted(self._all_categories, key=lambda c: c["name"])
        cats_ws = wb.create_sheet("Categorias")
        cats_ws.append(["id", "nome"])
        for c in categories:
            cats_ws.append([c["id"], c["name"]])

        if categories and "category" in fieldnames:
            col_letter = ws.cell(row=1, column=fieldnames.index("category") + 1).column_letter
            dv = DataValidation(type="list", formula1=f"=Categorias!$A$2:$A${len(categories) + 1}", allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}{len(self._mobs) + 1}")

        wb.save(path)
        logger.info("Mobs exportados para %s (%d linha(s))", path, len(self._mobs))

    def _export_xlsx_blank_template(self):
        """The Excel Importar card's "Template" button — a blank starting
        point to hand-fill (headers + one example row), not the user's
        current mobs. Distinct from _export_xlsx (the Exportar menu),
        which writes the real data instead of a placeholder row."""
        path, _selected = QFileDialog.getSaveFileName(
            self, "Baixar Template de Mobs (Excel)", "mobs_template.xlsx", "Planilha Excel (*.xlsx)",
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation

        fieldnames = list(_MOB_TEMPLATE_FIELDS.keys())
        example = dict(_MOB_TEMPLATE_FIELDS)
        example["name"] = "Novo Mob"
        wb = Workbook()
        ws = wb.active
        ws.title = "Mobs"
        ws.append(fieldnames)
        ws.append([example.get(k, "") for k in fieldnames])

        categories = sorted(self._all_categories, key=lambda c: c["name"])
        cats_ws = wb.create_sheet("Categorias")
        cats_ws.append(["id", "nome"])
        for c in categories:
            cats_ws.append([c["id"], c["name"]])

        if categories and "category" in fieldnames:
            col_letter = ws.cell(row=1, column=fieldnames.index("category") + 1).column_letter
            dv = DataValidation(type="list", formula1=f"=Categorias!$A$2:$A${len(categories) + 1}", allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}1000")

        wb.save(path)
        logger.info("Template Excel de mobs baixado: %s", path)

    def _import_mob_dicts(self, data) -> int:
        """Shared by all 3 Importar cards — the Excel card's staged file
        (any of json/csv/xlsx, see _on_file_dropped) and the JSON/CSV
        cards' Aplicar (see _build_text_import_card) alike each just
        parse their input into a list[dict] shaped like a mob row and
        hand it here for validation+create."""
        if not isinstance(data, list):
            logger.warning("Arquivo de importação inválido: esperada uma lista de mobs.")
            return 0
        known_columns = self._known_mob_columns()
        imported = 0
        image_matches = 0
        for entry in data:
            if not isinstance(entry, dict) or not entry.get("name"):
                continue
            fields = {k: v for k, v in entry.items()
                      if k not in ("id", "created_at", "updated_at") and (known_columns is None or k in known_columns)}
            # Importing/creating while browsing inside a folder files every
            # mob under that folder directly — like dropping files into a
            # directory in a file explorer — instead of whatever category
            # (or none) the source data happened to specify.
            if self._current_dir_id is not None:
                fields["category"] = self._current_dir_id
            mob_id = self._uow.mobs.create(**fields)
            imported += 1
            # A folder staged in the "Imagens (pasta)" card (even without
            # clicking its own Aplicar) auto-matches brand-new mobs too —
            # so creating mobs via JSON/CSV/Excel with a folder already
            # picked attaches images in the same step, no second pass
            # needed (see panel.py's _staged_image_files).
            if not fields.get("image_path") and self._staged_image_files:
                key = _normalize_mob_name(fields.get("name", ""))
                source_path = self._staged_image_files.get(key)
                if source_path:
                    try:
                        rel_path = import_asset(self._project_dir, source_path, "assets/mobs", mob_id)
                        self._uow.mobs.update(mob_id, image_path=rel_path)
                        image_matches += 1
                    except Exception:
                        logger.exception("Falha ao atribuir imagem ao mob recém-criado %s (%s)", fields.get("name", ""), source_path)
        logger.info(
            "Mobs importados: %d (categoria=%s, %d imagem(ns) auto-atribuída(s))",
            imported, self._current_dir_id, image_matches,
        )
        return imported

    def _known_mob_columns(self) -> set[str] | None:
        """The real mobs columns via PRAGMA, not `self._mobs[0].keys()` —
        that fallback silently disabled column filtering (returned None)
        whenever the project had zero mobs yet, exactly the situation a
        first-ever CSV/Excel import (far more likely than JSON to carry
        stray hand-edited columns) would hit."""
        if not self._uow:
            return None
        try:
            return set(self._uow.mobs.db.table_columns(self._uow.mobs.TABLE))
        except Exception:
            logger.exception("Falha ao introspectar colunas de mobs; import sem filtro de colunas.")
            return None

    # ─── Format-specific readers (file -> list[dict]) ───
    # resistances/drops_json are already JSON-encoded TEXT columns in the
    # DB — CSV/Excel just carry that same string verbatim in one cell, same
    # as JSON, so nothing special is needed to pass them through.

    def _read_json(self, path: str) -> list:
        return read_json(path)

    def _read_csv(self, path: str) -> list[dict]:
        return read_csv(path)

    @staticmethod
    def _normalize_blank_cells(row: dict) -> dict:
        return normalize_blank_cells(row)

    def _read_xlsx(self, path: str) -> list[dict]:
        return read_xlsx(path)
